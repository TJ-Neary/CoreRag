"""
Tests for src/extractor.py

Tests text extraction from various file types.
"""

import os
import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

from src.extractor import (
    _AUDIO_EXTENSIONS,
    _IMAGE_EXTENSIONS,
    _PDF_TEXT_THRESHOLD,
    _SPREADSHEET_EXTENSIONS,
    _VIDEO_EXTENSIONS,
    _extract_pdf,
    _extract_xlsx,
    extract_text,
)


class TestExtractText:
    """Tests for the main extract_text function."""

    def test_extract_txt_file(self, tmp_path):
        """Test extraction from plain text file."""
        txt_file = tmp_path / "test.txt"
        txt_file.write_text("Hello, this is a test document.")

        result = extract_text(txt_file)

        assert result == "Hello, this is a test document."

    def test_extract_md_file(self, tmp_path):
        """Test extraction from markdown file."""
        md_file = tmp_path / "test.md"
        md_file.write_text("# Heading\n\nSome **bold** text.")

        result = extract_text(md_file)

        assert "# Heading" in result
        assert "**bold**" in result

    def test_extract_json_file(self, tmp_path):
        """Test extraction from JSON file."""
        json_file = tmp_path / "test.json"
        json_file.write_text('{"key": "value", "nested": {"a": 1}}')

        result = extract_text(json_file)

        assert '"key": "value"' in result

    def test_extract_yaml_file(self, tmp_path):
        """Test extraction from YAML file."""
        yaml_file = tmp_path / "test.yaml"
        yaml_file.write_text("key: value\nlist:\n  - item1\n  - item2")

        result = extract_text(yaml_file)

        assert "key: value" in result

    def test_extract_csv_file(self, tmp_path):
        """Test extraction from CSV file."""
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("name,age,city\nAlice,30,NYC\nBob,25,LA")

        result = extract_text(csv_file)

        assert "name,age,city" in result
        assert "Alice,30,NYC" in result

    def test_extract_log_file(self, tmp_path):
        """Test extraction from log file."""
        log_file = tmp_path / "test.log"
        log_file.write_text(
            "2024-01-01 INFO: Application started\n2024-01-01 ERROR: Something failed"
        )

        result = extract_text(log_file)

        assert "INFO: Application started" in result
        assert "ERROR: Something failed" in result

    def test_extract_unsupported_extension(self, tmp_path):
        """Test that unsupported extensions return empty string."""
        unsupported = tmp_path / "test.xyz"
        unsupported.write_text("some content")

        result = extract_text(unsupported)

        assert result == ""

    def test_extract_nonexistent_file(self, tmp_path):
        """Test that nonexistent files return empty string."""
        result = extract_text(tmp_path / "does_not_exist.txt")

        assert result == ""

    def test_extract_empty_file(self, tmp_path):
        """Test extraction from empty file."""
        empty = tmp_path / "empty.txt"
        empty.write_text("")

        result = extract_text(empty)

        assert result == ""

    def test_extract_file_with_encoding_issues(self, tmp_path):
        """Test extraction handles encoding errors gracefully."""
        binary_file = tmp_path / "binary.txt"
        binary_file.write_bytes(b"Hello \xff\xfe World")

        result = extract_text(binary_file)

        # Should not raise, may have replacement characters
        assert "Hello" in result or "World" in result


class TestPDFExtraction:
    """Tests for PDF text extraction."""

    @patch("src.extractor.pypdf.PdfReader")
    def test_extract_text_based_pdf(self, mock_reader_class):
        """Test extraction from text-based PDF."""
        mock_reader = MagicMock()
        mock_page = MagicMock()
        mock_page.extract_text.return_value = "This is PDF content. " * 10  # > threshold
        mock_reader.pages = [mock_page]
        mock_reader_class.return_value = mock_reader

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            f.write(b"%PDF-1.4 fake pdf content")
            f.flush()
            try:
                result = _extract_pdf(Path(f.name))
                assert "This is PDF content" in result
            finally:
                os.unlink(f.name)

    @patch("src.extractor._extract_pdf_ocr")
    @patch("src.extractor.pypdf.PdfReader")
    def test_scanned_pdf_falls_back_to_ocr(self, mock_reader_class, mock_ocr):
        """Test that scanned PDFs (little text) fall back to OCR."""
        mock_reader = MagicMock()
        mock_page = MagicMock()
        mock_page.extract_text.return_value = "ab"  # Below threshold
        mock_reader.pages = [mock_page]
        mock_reader_class.return_value = mock_reader
        mock_ocr.return_value = "OCR extracted text"

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            f.write(b"%PDF-1.4 fake pdf content")
            f.flush()
            try:
                result = _extract_pdf(Path(f.name))
                assert mock_ocr.called
                assert result == "OCR extracted text"
            finally:
                os.unlink(f.name)


class TestImageExtraction:
    """Tests for image OCR extraction."""

    @patch("src.extractor._extract_image_ocr")
    def test_png_uses_ocr(self, mock_ocr):
        """Test PNG files use OCR extraction."""
        mock_ocr.return_value = "Text from image"

        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
            f.write(b"fake png content")
            f.flush()
            try:
                extract_text(Path(f.name))
                mock_ocr.assert_called_once()
            finally:
                os.unlink(f.name)

    def test_image_extensions_defined(self):
        """Test that expected image extensions are configured."""
        expected = {".png", ".jpg", ".jpeg", ".tiff", ".webp", ".heic"}
        assert expected.issubset(_IMAGE_EXTENSIONS)


class TestAudioExtraction:
    """Tests for audio transcription."""

    def test_audio_extensions_defined(self):
        """Test that expected audio extensions are configured."""
        expected = {".mp3", ".wav", ".m4a", ".flac"}
        assert expected.issubset(_AUDIO_EXTENSIONS)

    @patch("src.extractor._extract_audio")
    def test_mp3_uses_transcription(self, mock_audio):
        """Test MP3 files use audio extraction."""
        mock_audio.return_value = "Transcribed audio"

        with tempfile.NamedTemporaryFile(suffix=".mp3", delete=False) as f:
            f.write(b"fake mp3 content")
            f.flush()
            try:
                extract_text(Path(f.name))
                mock_audio.assert_called_once()
            finally:
                os.unlink(f.name)


class TestVideoExtraction:
    """Tests for video processing."""

    def test_video_extensions_defined(self):
        """Test that expected video extensions are configured."""
        expected = {".mp4", ".mov", ".avi", ".mkv"}
        assert expected.issubset(_VIDEO_EXTENSIONS)

    @patch("src.extractor._extract_video")
    def test_mp4_uses_video_extraction(self, mock_video):
        """Test MP4 files use video extraction."""
        mock_video.return_value = "Video analysis text"

        with tempfile.NamedTemporaryFile(suffix=".mp4", delete=False) as f:
            f.write(b"fake mp4 content")
            f.flush()
            try:
                extract_text(Path(f.name))
                mock_video.assert_called_once()
            finally:
                os.unlink(f.name)


class TestXLSXExtraction:
    """Tests for XLSX spreadsheet extraction."""

    def test_spreadsheet_extensions_defined(self):
        """Test that expected spreadsheet extensions are configured."""
        expected = {".xlsx", ".xls", ".xlsm"}
        assert expected == _SPREADSHEET_EXTENSIONS

    def test_extract_xlsx_basic(self, tmp_path):
        """Test extraction from a basic XLSX file."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Name", "Age", "City"])
        ws.append(["Alice", 30, "Springfield"])
        ws.append(["Bob", 25, "Shelbyville"])
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(xlsx_path)

        result = _extract_xlsx(xlsx_path)

        assert "## Data" in result
        assert "Alice" in result
        assert "Bob" in result
        assert "Springfield" in result

    def test_extract_xlsx_multiple_sheets(self, tmp_path):
        """Test extraction from XLSX with multiple sheets."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["Header1", "Header2"])
        ws1.append(["A", "B"])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["Col1", "Col2"])
        ws2.append(["X", "Y"])

        xlsx_path = tmp_path / "multi.xlsx"
        wb.save(xlsx_path)

        result = _extract_xlsx(xlsx_path)

        assert "## Sheet1" in result
        assert "## Sheet2" in result
        assert "Header1" in result
        assert "Col1" in result

    def test_extract_xlsx_empty_sheet(self, tmp_path):
        """Test extraction skips empty sheets."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "HasData"
        ws1.append(["Data", "Here"])

        wb.create_sheet("Empty")

        xlsx_path = tmp_path / "empty_sheet.xlsx"
        wb.save(xlsx_path)

        result = _extract_xlsx(xlsx_path)

        assert "## HasData" in result
        assert "## Empty" not in result

    def test_extract_xlsx_via_dispatch(self, tmp_path):
        """Test that .xlsx files are routed through extract_text dispatcher."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Test", "Content"])
        xlsx_path = tmp_path / "dispatch.xlsx"
        wb.save(xlsx_path)

        result = extract_text(xlsx_path)

        assert "Test" in result
        assert "Content" in result

    def test_extract_xlsx_with_none_cells(self, tmp_path):
        """Test extraction handles None cells gracefully."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", None, "C"])
        ws.append([None, "B", None])
        xlsx_path = tmp_path / "nulls.xlsx"
        wb.save(xlsx_path)

        result = _extract_xlsx(xlsx_path)

        assert "A" in result
        assert "C" in result


class TestConstants:
    """Tests for module constants."""

    def test_pdf_text_threshold_reasonable(self):
        """Test PDF text threshold is reasonable."""
        assert 10 <= _PDF_TEXT_THRESHOLD <= 200
